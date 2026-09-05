import json
import os
import openpyxl
import pandas as pd

def inspect_workbook(filepath):
    print(f"\n--- Inspecting {filepath} ---")
    file_size_mb = os.path.getsize(filepath) / (1024 * 1024)
    wb = openpyxl.load_workbook(filepath, read_only=True)
    sheets = wb.sheetnames
    
    inventory = {
        "filepath": filepath,
        "file_size_mb": round(file_size_mb, 2),
        "sheets": []
    }
    
    for sheet_name in sheets:
        ws = wb[sheet_name]
        max_r = ws.max_row
        max_c = ws.max_column
        print(f"Sheet: {sheet_name}, Rows: {max_r}, Cols: {max_c}")
        
        # Read first 15 rows with pandas
        df_head = pd.read_excel(filepath, sheet_name=sheet_name, nrows=15, header=None)
        
        sheet_info = {
            "sheet_name": sheet_name,
            "max_rows": max_r,
            "max_cols": max_c,
            "sample_head": df_head.fillna("").values.tolist()
        }
        inventory["sheets"].append(sheet_info)
        
    wb.close()
    return inventory

if __name__ == "__main__":
    inv_2025 = inspect_workbook("raw/2025/GAA-2025.xlsx")
    inv_2026 = inspect_workbook("raw/2026/FY2026-GAA-Byobject.xlsx")
    
    os.makedirs("reports/validation", exist_ok=True)
    with open("reports/validation/inventory_raw.json", "w") as f:
        json.dump({"2025": inv_2025, "2026": inv_2026}, f, indent=2)
    print("\nSaved raw inventory to reports/validation/inventory_raw.json")
